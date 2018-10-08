from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from openpyxl import Workbook

# Loading Firefox profile
profile = FirefoxProfile("C:\\Users\\mjoys\\AppData\\Roaming\\Mozilla\\Firefox\\Profiles\\a8wg4ay3.default")
# Firefox web-driver
browser = webdriver.Firefox(firefox_profile=profile)
# Opening Link of facebook group
browser.get('https://www.facebook.com/groups/cookupsBD/')

# Wait 60 seconds for page to load
timeout = 60

try:
    WebDriverWait(browser, timeout).until(
        EC.visibility_of_element_located((By.XPATH, "//img[@class='_4on7 _3mk2 img']")))

except TimeoutException:
    browser.quit()

# find_elements_by_xpath returns an array of selenium objects.
post_divs = browser.find_elements_by_xpath("//div[@class='_4-u2 mbm _4mrt _5jmm _5pat _5v3q _4-u8']")
post_count = 10
names, times, post_titles, cook_locations, prices, descriptions, offers, images = ([] for i in range(8))
for items in post_divs:
    try:
        price_tag = items.find_element_by_class_name('_l57')
        print("price_tag", price_tag)
        if price_tag:
            prices.append(price_tag.text)
            names.append(items.find_element_by_class_name('profileLink').text)
            times.append(items.find_element_by_class_name('timestampContent').text)
            post_titles.append(items.find_element_by_class_name('_l53').text)
            cook_locations.append(items.find_element_by_class_name('_l58').text)
            descriptions.append(items.find_element_by_class_name('text_exposed_root').text)
    except Exception as e:
        pass
headers = ['Cook Name', 'Time of Post ', 'Title of Post', 'Cook Location', 'Price', 'Description', 'Order Link',
           'Image']
wb = Workbook()
ws = wb.active
for i in range(1, 9):
    ws.cell(row=1, column=i).value = headers[i - 1]
c = 1
data_col = 1
for name, time, post_title, cook_location, price, description in zip(names[:post_count], times[:post_count],
                                                                     post_titles[:post_count],
                                                                     cook_locations[:post_count], prices[:post_count],
                                                                     descriptions[:post_count]):
    c = c + 1
    ws.cell(row=c, column=data_col).value = name
    data_col = data_col + 1
    ws.cell(row=c, column=data_col).value = time
    data_col = data_col + 1
    ws.cell(row=c, column=data_col).value = post_title
    data_col = data_col + 1
    ws.cell(row=c, column=data_col).value = cook_location
    data_col = data_col + 1
    ws.cell(row=c, column=data_col).value = price
    data_col = data_col + 1
    ws.cell(row=c, column=data_col).value = description
    data_col = data_col + 1
    ws.cell(row=c, column=data_col).value = ''
    data_col = data_col + 1
    ws.cell(row=c, column=data_col).value = ''
    data_col = 1
wb.save('cookups.xlsx')
